
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def create_biodym_template_v2():
    """
    Creates a new, structured Excel template for BioDYM v2.
    """
    wb = Workbook()

    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # --- 00_About ---
    ws_readme: Worksheet = wb.create_sheet("00_ReadMe")
    ws_readme.cell(row=1, column=1, value="BioDYM MFA Template v2.0")
    ws_readme.cell(row=2, column=1, value="This template provides a standardized and hierarchical structure for defining MFA models.")
    
    ws_codelists: Worksheet = wb.create_sheet("00_Codelists")
    ws_codelists.cell(row=1, column=1, value="Distribution Types")
    ws_codelists.cell(row=2, column=1, value="Normal")
    ws_codelists.cell(row=3, column=1, value="Uniform")
    ws_codelists.cell(row=4, column=1, value="Triangular")
    ws_codelists.cell(row=5, column=1, value="Lognormal")


    # --- 01_System_Definition ---
    ws_processes: Worksheet = wb.create_sheet("01_Processes")
    ws_processes.append(["Process_ID", "Name", "Model_Type", "Description"])
    ws_processes.append(["P01_Example", "Example Process", "Splitter", "A descriptive name for the process."])

    ws_flows: Worksheet = wb.create_sheet("01_Flows")
    ws_flows.append(["Flow_ID", "P_Start_ID", "P_End_ID", "Description"])
    ws_flows.append(["F01_02_Example", "P01_Example", "P02_Another", "A descriptive name for the flow."])

    # --- 02_Model_Parameters ---
    ws_flow_comp: Worksheet = wb.create_sheet("02_Flow_Composition")
    ws_flow_comp.append(["Flow_ID", "Element_ID", "Value", "Unit"])
    ws_flow_comp.append(["F01_02_Example", "DM", 0.8, "fraction"])

    ws_tc: Worksheet = wb.create_sheet("02_TC_Parameters")
    ws_tc.append(["Parameter_ID", "Description", "Value", "Unit"])
    ws_tc.append(["TC_P01_Splitter_01", "Example TC for P01", 0.5, "fraction"])

    ws_dsm: Worksheet = wb.create_sheet("02_DSM_Parameters")
    ws_dsm.append(["Parameter_ID", "Process_ID", "Parameter_Name", "Value", "Unit"])
    ws_dsm.append(["DSM_P03_Lifetime_01", "P03_Use_Phase", "Mean", 10, "years"])

    ws_fomp: Worksheet = wb.create_sheet("02_FOMP_Parameters")
    ws_fomp.append(["Parameter_ID", "Process_ID", "Parameter_Name", "Value", "Unit"])
    ws_fomp.append(["FOMP_P04_Decay_Labile", "P04_Decay", "Decay_Rate_Labile", 0.2, "1/year"])

    # --- 03_Uncertainty_Analysis ---
    ws_unc: Worksheet = wb.create_sheet("03_Uncertainty_Parameters")
    ws_unc.append(["Uncertainty_ID", "Parameter_ID", "Distribution_Type", "Mean_Value", "Std_Dev", "Min", "Max"])
    ws_unc.append(["UNC_TC_P01_Splitter_01_Normal", "TC_P01_Splitter_01", "Normal", 0.5, 0.05, "", ""])

    # --- 04_Scenario_Management ---
    ws_groups: Worksheet = wb.create_sheet("04_Parameter_Groups")
    ws_groups.append(["Group_ID", "Parameter_ID", "Description"])
    ws_groups.append(["High_Efficiency", "TC_P01_Splitter_01", "Sets the main splitter to a high value."])

    ws_scenarios: Worksheet = wb.create_sheet("04_Scenarios")
    ws_scenarios.append(["Scenario_ID", "Description", "Groups_to_Include"])
    ws_scenarios.append(["High_Tech_Future", "A scenario assuming high technology adoption.", "High_Efficiency"])

    # --- 05_Visualization ---
    ws_layout_p: Worksheet = wb.create_sheet("05_Layout_Processes")
    ws_layout_p.append(["Process_ID", "x", "y", "Color"])
    ws_layout_p.append(["P01_Example", 100, 100, "#FF0000"])

    ws_layout_f: Worksheet = wb.create_sheet("05_Layout_Flows")
    ws_layout_f.append(["Flow_ID", "Color", "Style"])
    ws_layout_f.append(["F01_02_Example", "#0000FF", "solid"])
    
    # Save the workbook
    output_path = "BioDYM_Template_V2.xlsx"
    wb.save(output_path)
    print(f"Successfully created template: {output_path}")

if __name__ == "__main__":
    create_biodym_template_v2()
