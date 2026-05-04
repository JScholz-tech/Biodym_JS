# -*- coding: utf-8 -*-
"""
Simplified Sankey Diagram Plotting Module for BioDYM.

Pure automatic positioning using Plotly's built-in layout engine.
For manual positioning control, use enhanced_sankey.py instead.

Key features:
- Interactive single-element Sankey diagrams with year/process/flow filtering
- Element-agnostic coloring (works with any element set)
- Multi-element vertical stacked visualizations
- Automatic node positioning via Plotly's 'snap' arrangement
- Publication-quality styling

Author: BioDYM Development Team
Date: 2025-11-10 (Clean - Automatic Only)
"""

import plotly.graph_objects as go
from ipywidgets import (
    FloatSlider,
    IntSlider,
    HBox,
    VBox,
    HTML,
    Layout,
    Dropdown,
    SelectMultiple,
    interactive,
)
from IPython.display import display
from typing import Optional

from .themes import (
    get_process_color,
    detect_biodym_process_type,
    get_stock_color,
    BIOYM_COLORS,
    FONT_FAMILY,
)
from . import sankey_config
from .dynamic_colors import ElementColorManager


def _wrap_label_text(text, max_chars):
    """Wrap long text labels with HTML line breaks.

    Respects existing <br> tags and wraps at word boundaries.

    Parameters
    ----------
    text : str
        The text to wrap
    max_chars : int
        Maximum characters per line (0 = no wrapping)

    Returns
    -------
    str
        Text with <br> tags inserted at appropriate points
    """
    if not sankey_config.NODE_LABEL_WRAP or max_chars <= 0:
        return text

    # If text already has manual line breaks, keep them
    if "<br>" in text.lower():
        return text

    # If text is short enough, no wrapping needed
    if len(text) <= max_chars:
        return text

    # Wrap at word boundaries (including underscores and hyphens)
    # Split on spaces, underscores, and hyphens while preserving the separators
    import re
    # Split but keep the separators
    parts = re.split(r'(\s+|_|-)', text)
    # Filter out empty strings
    parts = [p for p in parts if p]

    lines = []
    current_line = []
    current_length = 0

    for part in parts:
        part_length = len(part)

        # Check if adding this part would exceed max_chars
        if current_length + part_length <= max_chars:
            current_line.append(part)
            current_length += part_length
        else:
            # Start new line if current line is not empty
            if current_line:
                lines.append("".join(current_line))
                current_line = []
                current_length = 0

            # If the part itself is longer than max_chars, add it anyway
            # (better to have one long line than break in the middle of a word)
            if part_length > max_chars:
                if current_line:
                    lines.append("".join(current_line))
                    current_line = []
                    current_length = 0
                lines.append(part)
            else:
                current_line.append(part)
                current_length = part_length

    # Add last line
    if current_line:
        lines.append("".join(current_line))

    return "<br>".join(lines)


def _prepare_sankey_data(
    filtered_processes,
    flows_data,
    mfa_system_results,
    dsm_params,
    fomp_params,
    node_pad,
    node_thickness,
):
    """Prepare node and link data for Plotly Sankey.

    Uses Plotly's automatic positioning exclusively - no manual overrides.

    Parameters
    ----------
    filtered_processes : list of odym.Process
        Processes to display
    flows_data : list of tuples
        List of (flow_obj, value) tuples for the current year/element
    mfa_system_results : odym.MFAsystem
        MFA system results
    dsm_params : dict or None
        DSM parameters
    fomp_params : dict or None
        FOMP parameters
    node_pad : int
        Padding between nodes in pixels
    node_thickness : int
        Thickness of node bars in pixels

    Returns
    -------
    dict, dict
        Node data dict and link data dict for go.Sankey
    """
    # Build process ID to index mapping
    process_id_map = {p.ID: i for i, p in enumerate(filtered_processes)}

    # Prepare node labels with automatic wrapping
    node_labels = [
        _wrap_label_text(p.Name, sankey_config.NODE_LABEL_MAX_CHARS)
        for p in filtered_processes
    ]

    # Determine node colors based on process type
    node_colors = []
    for p in filtered_processes:
        has_stocks = f"S_{p.ID}" in mfa_system_results.StockDict
        is_dsm = dsm_params and p.ID in dsm_params
        is_fomp = fomp_params and p.ID in fomp_params

        if is_dsm or is_fomp:
            process_type = detect_biodym_process_type(
                p.ID, dsm_params=dsm_params, fomp_params=fomp_params
            )
            node_colors.append(get_process_color(process_type))
        elif has_stocks:
            node_colors.append(get_stock_color())
        else:
            node_colors.append(get_process_color("regular"))

    # Build node dict - Plotly handles ALL positioning automatically
    node_dict = {
        "label": node_labels,
        "color": node_colors,
        "pad": node_pad,
        "thickness": node_thickness,
        "line": {"color": "black", "width": 0.5},
    }
    # Note: Font styling is applied at layout level, not on individual nodes

    # Prepare link data
    if not flows_data:
        return node_dict, {"source": [], "target": [], "value": []}

    link_sources = []
    link_targets = []
    link_values = []
    link_customdata = []

    for flow, value in flows_data:
        if flow.P_Start in process_id_map and flow.P_End in process_id_map:
            link_sources.append(process_id_map[flow.P_Start])
            link_targets.append(process_id_map[flow.P_End])
            link_values.append(value)
            link_customdata.append(getattr(flow, "DescriptiveName", flow.Name))

    link_dict = {
        "source": link_sources,
        "target": link_targets,
        "value": link_values,
        "customdata": link_customdata,
        "hovertemplate": "Flow: %{customdata}<br />Source: %{source.label}<br />Target: %{target.label}<br />Value: %{value}<extra></extra>",
    }

    # Add arrows to flows if enabled
    if sankey_config.ENABLE_FLOW_ARROWS:
        link_dict["arrowlen"] = sankey_config.FLOW_ARROW_LENGTH

    return node_dict, link_dict


def _create_sankey_widgets(
    all_process_names, all_flow_names, time_items, element_items, max_flow_value
):
    """Create UI widgets for interactive Sankey diagram."""
    year_slider = IntSlider(
        min=time_items[0],
        max=time_items[-1],
        step=1,
        value=time_items[0],
        description="Year:",
        style={"description_width": "80px"},
        layout=Layout(width="400px"),
    )

    element_dropdown = Dropdown(
        options=element_items,
        value=element_items[0],
        description="Element:",
        style={"description_width": "80px"},
        layout=Layout(width="200px"),
    )

    process_selector = SelectMultiple(
        options=all_process_names,
        value=tuple(all_process_names),
        description="Processes:",
        style={"description_width": "100px"},
        layout=Layout(width="400px", height="120px"),
    )

    flow_selector = SelectMultiple(
        options=all_flow_names,
        value=tuple(all_flow_names),
        description="Flows:",
        style={"description_width": "100px"},
        layout=Layout(width="400px", height="120px"),
    )

    threshold_slider = FloatSlider(
        min=0,
        max=max_flow_value,
        step=max_flow_value / 100 if max_flow_value > 0 else 0.01,
        value=0,
        description="Min Flow:",
        continuous_update=False,
        readout_format=".2f",
        style={"description_width": "80px"},
        layout=Layout(width="400px"),
    )

    return (
        year_slider,
        element_dropdown,
        process_selector,
        flow_selector,
        threshold_slider,
    )


def _create_sankey_legend(element_items, color_manager):
    """Create HTML legend for Sankey diagram."""
    flow_legend_items = []
    for element in element_items:
        element_color = color_manager.get_element_color(element.lower())
        flow_legend_items.append(f"""
                <div style="display: flex; align-items: center;">
                    <div style="width: 20px; height: 20px; background-color: {element_color}; margin-right: 5px;"></div>
                    <span>{element.upper()}</span>
                </div>""")

    flow_legend_html = "\n".join(flow_legend_items)

    legend_html = f"""
    <div style="margin: 10px; padding: 10px; border: 1px solid #ccc; border-radius: 5px; background-color: #f9f9f9;">
        <h4 style="margin: 0 0 10px 0;">Legend</h4>
        <div style="margin-bottom: 10px;">
            <strong>Processes / Stocks:</strong><br>
            <div style="display: flex; flex-wrap: wrap; gap: 15px; margin-top: 5px;">
                <div style="display: flex; align-items: center;">
                    <div style="width: 20px; height: 20px; background-color: {get_process_color("regular")}; margin-right: 5px;"></div>
                    <span>Regular processes</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <div style="width: 20px; height: 20px; background-color: {get_stock_color()}; margin-right: 5px;"></div>
                    <span>Stocks</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <div style="width: 20px; height: 20px; background-color: {get_process_color("dsm")}; margin-right: 5px;"></div>
                    <span>DSM Process</span>
                </div>
                <div style="display: flex; align-items: center;">
                    <div style="width: 20px; height: 20px; background-color: {get_process_color("fomp")}; margin-right: 5px;"></div>
                    <span>FOMP Process</span>
                </div>
            </div>
        </div>
        <div>
            <strong>Flows:</strong><br>
            <div style="display: flex; flex-wrap: wrap; gap: 15px; margin-top: 5px;">
                {flow_legend_html}
            </div>
        </div>
    </div>
    """
    return HTML(value=legend_html)


def plot_interactive_sankey(
    mfa_system_results,
    dsm_params=None,
    fomp_params=None,
    color_manager: Optional[ElementColorManager] = None,
    width=sankey_config.WINDOW_WIDTH,
    height=sankey_config.WINDOW_HEIGHT,
    node_pad=sankey_config.NODE_SPACING,
    node_thickness=sankey_config.NODE_THICKNESS,
):
    """Display an interactive Sankey diagram for MFA results.

    Uses Plotly's automatic node positioning exclusively.
    For manual positioning control, use enhanced_sankey module instead.

    Parameters
    ----------
    mfa_system_results : odym.MFAsystem
        Solved MFA system with calculated data
    dsm_params : dict, optional
        DSM parameters for process coloring
    fomp_params : dict, optional
        FOMP parameters for process coloring
    color_manager : ElementColorManager, optional
        Color manager for elements (auto-created if None)
    width : int, optional
        Diagram width in pixels
    height : int, optional
        Diagram height in pixels
    node_pad : int, optional
        Padding between nodes in pixels
    node_thickness : int, optional
        Thickness of node bars in pixels

    Examples
    --------
    >>> plot_interactive_sankey(mfa_results_baseline)

    >>> plot_interactive_sankey(
    ...     mfa_results_baseline,
    ...     width=4000,
    ...     height=1500,
    ...     node_pad=30
    ... )
    """
    # Extract data from MFA system
    all_process_names = [p.Name for p in mfa_system_results.ProcessList]

    # Get flow descriptions if available, otherwise use Flow IDs
    flow_descriptions = getattr(mfa_system_results, "_flow_descriptions", {})
    all_flow_names = [
        flow_descriptions.get(f.Name, f.Name)
        for f in mfa_system_results.FlowDict.values()
    ]
    all_flows = list(mfa_system_results.FlowDict.values())
    time_items = mfa_system_results.IndexTable.Classification["Time"].Items
    element_items = mfa_system_results.Elements

    # Create color manager if needed
    if color_manager is None:
        color_manager = ElementColorManager([e.lower() for e in element_items])

    max_flow_value = (
        max(f.Values.max() for f in all_flows if f.Values is not None)
        if all_flows
        else 1
    )

    # Create figure widget - no arrangement constraint for maximum flexibility
    fig = go.FigureWidget(data=[go.Sankey(node={}, link={})])

    def update_sankey(year, element, processes_to_show, flows_to_show, min_flow_value):
        """Update callback for widget changes."""
        if not processes_to_show or not flows_to_show:
            with fig.batch_update():
                fig.data[0].node.label = []
            return

        # Filter processes
        filtered_processes = [
            p for p in mfa_system_results.ProcessList if p.Name in processes_to_show
        ]
        process_ids = [p.ID for p in filtered_processes]
        process_id_set = set(process_ids)

        # Get year and element indices
        year_index = time_items.index(year)
        element_index = element_items.index(element)

        # Filter flows
        # Create mapping from descriptive names back to Flow IDs
        flows_to_show_ids = set()
        for flow_id, desc_name in flow_descriptions.items():
            if desc_name in flows_to_show:
                flows_to_show_ids.add(flow_id)
        # Also add any Flow IDs that are directly in flows_to_show (backward compatibility)
        flows_to_show_ids.update(flows_to_show)

        flows_data = []
        for flow in all_flows:
            if (
                flow.Name in flows_to_show_ids
                and flow.P_Start in process_id_set
                and flow.P_End in process_id_set
            ):
                flow_value = flow.Values[year_index, element_index]
                if flow_value >= min_flow_value:
                    flows_data.append((flow, flow_value))

        # Prepare node and link data (Plotly handles all positioning)
        node_dict, link_dict = _prepare_sankey_data(
            filtered_processes,
            flows_data,
            mfa_system_results,
            dsm_params,
            fomp_params,
            node_pad,
            node_thickness,
        )

        # Set link colors based on element
        element_color = color_manager.get_element_color(element.lower())
        if link_dict["value"]:
            link_dict["color"] = [element_color] * len(link_dict["value"])

        # Update figure
        with fig.batch_update():
            fig.data[0].node = node_dict
            fig.data[0].link = link_dict

            # Update layout
            # Format element name according to config
            if sankey_config.TITLE_ELEMENT_FORMAT == "upper":
                formatted_element = element.upper()
            elif sankey_config.TITLE_ELEMENT_FORMAT == "lower":
                formatted_element = element.lower()
            else:  # "title" is default
                formatted_element = element.title()

            title_text = sankey_config.TITLE_TEMPLATE.format(
                element=formatted_element, year=year
            )
            layout_config = {
                "title": {
                    "text": title_text,
                    "x": 0.5,
                    "xanchor": "center",
                    "font": {
                        "family": FONT_FAMILY,
                        "size": sankey_config.FONT_SIZE_TITLE,
                        "color": BIOYM_COLORS["dark"],
                    },
                },
                "font": {
                    "family": FONT_FAMILY,
                    "size": sankey_config.FONT_SIZE_LABELS,
                },
                "width": width,
                "height": height,
                "paper_bgcolor": sankey_config.BACKGROUND_COLOR,
                "plot_bgcolor": sankey_config.GRID_COLOR,
                "margin": dict(
                    l=width * sankey_config.PADDING_FACTOR,
                    r=width * sankey_config.PADDING_FACTOR,
                    t=height * sankey_config.PADDING_FACTOR,
                    b=height * sankey_config.PADDING_FACTOR,
                ),
            }

            # Add vertical grid lines if enabled
            if (
                sankey_config.ENABLE_GRID
                and sankey_config.GRID_TYPE == "vertical_lines"
            ):
                shapes = []
                for x_pos in sankey_config.GRID_VERTICAL_POSITIONS:
                    shapes.append(
                        dict(
                            type="line",
                            x0=x_pos,
                            x1=x_pos,
                            y0=0,
                            y1=1,
                            xref="paper",
                            yref="paper",
                            line=dict(
                                color=sankey_config.GRID_LINE_COLOR,
                                width=sankey_config.GRID_LINE_WIDTH,
                                dash=sankey_config.GRID_LINE_DASH,
                            ),
                            opacity=sankey_config.GRID_LINE_OPACITY,
                            layer="below",
                        )
                    )
                layout_config["shapes"] = shapes

            fig.update_layout(**layout_config)

    # Create widgets
    year_slider, element_dropdown, process_selector, flow_selector, threshold_slider = (
        _create_sankey_widgets(
            all_process_names, all_flow_names, time_items, element_items, max_flow_value
        )
    )
    legend_widget = _create_sankey_legend(element_items, color_manager)

    # Layout widgets (matches enhanced_sankey structure)
    ui = VBox(
        [
            HBox([year_slider, element_dropdown]),
            HBox([process_selector, flow_selector]),
            HBox([threshold_slider]),
            legend_widget,
        ]
    )

    # Connect widgets
    interactive(
        update_sankey,
        year=year_slider,
        element=element_dropdown,
        processes_to_show=process_selector,
        flows_to_show=flow_selector,
        min_flow_value=threshold_slider,
    )

    # Display (matches enhanced_sankey: VBox with controls and figure)
    display(VBox([ui, fig]))

    # Initial draw - explicitly trigger update with default values
    update_sankey(
        year_slider.value,
        element_dropdown.value,
        process_selector.value,
        flow_selector.value,
        threshold_slider.value,
    )


def plot_element_multiplot_sankey(
    mfa_system_results,
    dsm_params=None,
    fomp_params=None,
    elements_to_plot=None,
    subplot_height=sankey_config.WINDOW_HEIGHT,
    subplot_width=sankey_config.WINDOW_WIDTH,
    node_pad=sankey_config.NODE_SPACING,
    node_thickness=sankey_config.NODE_THICKNESS,
    color_manager: Optional[ElementColorManager] = None,
):
    """Display multiple Sankey diagrams stacked vertically (one per element).

    All subplots share time slider, process selector, and min flow filter.
    Uses automatic positioning exclusively.

    Parameters
    ----------
    mfa_system_results : odym.MFAsystem
        Solved MFA system
    dsm_params : dict, optional
        DSM parameters
    fomp_params : dict, optional
        FOMP parameters
    elements_to_plot : list of str, optional
        Elements to plot (default: all)
    subplot_height : int, optional
        Height per subplot in pixels
    subplot_width : int, optional
        Width per subplot in pixels
    node_pad : int, optional
        Padding between nodes in pixels
    node_thickness : int, optional
        Thickness of node bars in pixels
    color_manager : ElementColorManager, optional
        Color manager (auto-created if None)
    """
    # Extract data
    all_process_names = [p.Name for p in mfa_system_results.ProcessList]

    # Get flow descriptions if available, otherwise use Flow IDs
    flow_descriptions = getattr(mfa_system_results, "_flow_descriptions", {})
    all_flow_names = [
        flow_descriptions.get(f.Name, f.Name)
        for f in mfa_system_results.FlowDict.values()
    ]
    all_flows = list(mfa_system_results.FlowDict.values())
    time_items = mfa_system_results.IndexTable.Classification["Time"].Items
    element_items = mfa_system_results.Elements

    # Create color manager
    if color_manager is None:
        color_manager = ElementColorManager([e.lower() for e in element_items])

    # Determine elements to plot
    if elements_to_plot is None:
        elements_to_plot = element_items
    else:
        elements_to_plot = [e for e in elements_to_plot if e in element_items]

    max_flow_value = (
        max(f.Values.max() for f in all_flows if f.Values is not None)
        if all_flows
        else 1
    )

    # Create one figure per element
    figures = []
    for elem in elements_to_plot:
        fig = go.FigureWidget(data=[go.Sankey(node={}, link={})])

        # Build layout config
        # Format element name according to config
        if sankey_config.SUBPLOT_ELEMENT_FORMAT == "upper":
            formatted_elem = elem.upper()
        elif sankey_config.SUBPLOT_ELEMENT_FORMAT == "lower":
            formatted_elem = elem.lower()
        else:  # "title" is default
            formatted_elem = elem.title()

        subplot_title = sankey_config.SUBPLOT_TITLE_TEMPLATE.format(
            element=formatted_elem
        )

        layout_config = dict(
            height=subplot_height,
            width=subplot_width,
            title_text=subplot_title,
            font=dict(
                family=FONT_FAMILY,
                size=sankey_config.FONT_SIZE_LABELS,
            ),
            title_font_size=sankey_config.FONT_SIZE_SUBPLOT,
            margin=dict(
                l=subplot_width * sankey_config.PADDING_FACTOR,
                r=subplot_width * sankey_config.PADDING_FACTOR,
                t=subplot_height * sankey_config.PADDING_FACTOR,
                b=subplot_height * sankey_config.PADDING_FACTOR,
            ),
            paper_bgcolor=sankey_config.BACKGROUND_COLOR,
            plot_bgcolor=sankey_config.GRID_COLOR,
        )

        # Add vertical grid lines if enabled
        if sankey_config.ENABLE_GRID and sankey_config.GRID_TYPE == "vertical_lines":
            shapes = []
            for x_pos in sankey_config.GRID_VERTICAL_POSITIONS:
                shapes.append(
                    dict(
                        type="line",
                        x0=x_pos,
                        x1=x_pos,
                        y0=0,
                        y1=1,
                        xref="paper",
                        yref="paper",
                        line=dict(
                            color=sankey_config.GRID_LINE_COLOR,
                            width=sankey_config.GRID_LINE_WIDTH,
                            dash=sankey_config.GRID_LINE_DASH,
                        ),
                        opacity=sankey_config.GRID_LINE_OPACITY,
                        layer="below",
                    )
                )
            layout_config["shapes"] = shapes

        fig.update_layout(**layout_config)
        figures.append(fig)

    def update_all_sankeys(year, processes_to_show, flows_to_show, min_flow_value):
        """Update all element subplots."""
        if not processes_to_show or not flows_to_show:
            for fig in figures:
                with fig.batch_update():
                    fig.data[0].node.label = []
            return

        # Filter processes
        filtered_processes = [
            p for p in mfa_system_results.ProcessList if p.Name in processes_to_show
        ]
        process_ids = [p.ID for p in filtered_processes]
        process_id_set = set(process_ids)
        year_index = time_items.index(year)

        # Create mapping from descriptive names back to Flow IDs
        flows_to_show_ids = set()
        for flow_id, desc_name in flow_descriptions.items():
            if desc_name in flows_to_show:
                flows_to_show_ids.add(flow_id)
        # Also add any Flow IDs that are directly in flows_to_show (backward compatibility)
        flows_to_show_ids.update(flows_to_show)

        # Update each figure
        for fig_idx, element in enumerate(elements_to_plot):
            element_index = element_items.index(element)

            # Filter flows for this element
            flows_data = []
            for flow in all_flows:
                if (
                    flow.Name in flows_to_show_ids
                    and flow.P_Start in process_id_set
                    and flow.P_End in process_id_set
                ):
                    flow_value = flow.Values[year_index, element_index]
                    if flow_value >= min_flow_value:
                        flows_data.append((flow, flow_value))

            # Prepare data (Plotly handles all positioning)
            node_dict, link_dict = _prepare_sankey_data(
                filtered_processes,
                flows_data,
                mfa_system_results,
                dsm_params,
                fomp_params,
                node_pad,
                node_thickness,
            )

            # Set link colors
            element_color = color_manager.get_element_color(element.lower())
            if link_dict["value"]:
                link_dict["color"] = [element_color] * len(link_dict["value"])

            # Update figure
            with figures[fig_idx].batch_update():
                figures[fig_idx].data[0].node = node_dict
                figures[fig_idx].data[0].link = link_dict

    # Create widgets (no element dropdown for multiplot)
    year_slider, _, process_selector, flow_selector, threshold_slider = (
        _create_sankey_widgets(
            all_process_names, all_flow_names, time_items, element_items, max_flow_value
        )
    )

    ui = VBox(
        [HBox([year_slider, threshold_slider]), HBox([process_selector, flow_selector])]
    )

    # Connect widgets
    interactive(
        update_all_sankeys,
        year=year_slider,
        processes_to_show=process_selector,
        flows_to_show=flow_selector,
        min_flow_value=threshold_slider,
    )

    # Display
    display(ui)
    for fig in figures:
        display(fig)

    # Initial draw
    update_all_sankeys(
        year_slider.value,
        process_selector.value,
        flow_selector.value,
        threshold_slider.value,
    )


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------

def export_sankey_json(mfa_system_results, year, element, filepath,
                       dsm_params=None, fomp_params=None, min_flow=0.0):
    """Export Sankey data as a D3-compatible JSON file.

    The output follows the standard D3 Sankey JSON format (nodes + links),
    which is accepted by D3-based web viewers, any future Python Sankey library,
    and can be imported into e-Sankey via its Excel data import feature.

    Parameters
    ----------
    mfa_system_results : odym.MFAsystem
        Solved MFA system with calculated flow values.
    year : int
        Calendar year to export (e.g. 2025).
    element : str
        Element name to export (e.g. "material", "DM", "CC").
    filepath : str or Path
        Output file path (should end in .json).
    dsm_params : dict, optional
        DSM parameters (used to identify DSM processes for node coloring).
    fomp_params : dict, optional
        FOMP parameters (used to identify FOMP processes for node coloring).
    min_flow : float, optional
        Minimum flow value to include (default 0 = include all non-zero flows).

    Returns
    -------
    dict
        The exported data structure (also written to filepath).
    """
    import json
    import pathlib

    if dsm_params is None:
        dsm_params = {}
    if fomp_params is None:
        fomp_params = {}

    time_vector = mfa_system_results.Time_V
    if year not in time_vector:
        raise ValueError(f"Year {year} not in model time range {time_vector[0]}-{time_vector[-1]}.")
    year_idx = list(time_vector).index(year)

    elements = mfa_system_results.Elements
    if element not in elements:
        raise ValueError(f"Element '{element}' not in system elements: {elements}.")
    elem_idx = elements.index(element)

    process_list = mfa_system_results.ProcessList
    process_id_map = {p.ID: i for i, p in enumerate(process_list)}
    color_manager = ElementColorManager(elements)

    nodes = []
    for p in process_list:
        proc_type = detect_biodym_process_type(p.ID, dsm_params=dsm_params, fomp_params=fomp_params)
        color = get_process_color(proc_type)
        nodes.append({
            "id":    p.ID,
            "name":  p.Name,
            "color": color,
            "type":  proc_type,
        })

    link_color = color_manager.get_element_color(element)
    links = []
    for flow in mfa_system_results.FlowDict.values():
        value = float(flow.Values[year_idx, elem_idx])
        if value <= min_flow:
            continue
        if flow.P_Start not in process_id_map or flow.P_End not in process_id_map:
            continue
        links.append({
            "source": process_id_map[flow.P_Start],
            "target": process_id_map[flow.P_End],
            "value":  round(value, 6),
            "label":  getattr(flow, "DescriptiveName", flow.Name),
            "color":  link_color,
        })

    payload = {
        "metadata": {
            "year":    year,
            "element": element,
            "unit":    "Mg",
            "source":  "BioDYM",
        },
        "nodes": nodes,
        "links": links,
    }

    pathlib.Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"Sankey JSON exported: {filepath}  ({len(nodes)} nodes, {len(links)} links)")
    return payload


def export_sankey_html(mfa_system_results, year, element, filepath,
                       dsm_params=None, fomp_params=None, min_flow=0.0,
                       title=None):
    """Export an interactive standalone Sankey HTML file (no Python required to view).

    The output is a self-contained HTML file with the Plotly Sankey diagram
    embedded. It can be shared with collaborators, embedded in reports, or
    opened in any web browser without a Python environment.

    Parameters
    ----------
    mfa_system_results : odym.MFAsystem
        Solved MFA system with calculated flow values.
    year : int
        Calendar year to export.
    element : str
        Element name to export.
    filepath : str or Path
        Output file path (should end in .html).
    dsm_params : dict, optional
        DSM parameters for node coloring.
    fomp_params : dict, optional
        FOMP parameters for node coloring.
    min_flow : float, optional
        Minimum flow value to include (default 0).
    title : str, optional
        Figure title. Defaults to "BioDYM Sankey - {element} ({year})".

    Returns
    -------
    plotly.graph_objects.Figure
        The Plotly figure (also written to filepath).
    """
    import pathlib

    if dsm_params is None:
        dsm_params = {}
    if fomp_params is None:
        fomp_params = {}

    time_vector = mfa_system_results.Time_V
    if year not in time_vector:
        raise ValueError(f"Year {year} not in model time range {time_vector[0]}-{time_vector[-1]}.")
    year_idx = list(time_vector).index(year)

    elements = mfa_system_results.Elements
    if element not in elements:
        raise ValueError(f"Element '{element}' not in system elements: {elements}.")
    elem_idx = elements.index(element)

    process_list = mfa_system_results.ProcessList
    process_id_map = {p.ID: i for i, p in enumerate(process_list)}
    color_manager = ElementColorManager(elements)
    link_color = color_manager.get_element_color(element)

    node_labels = [p.Name for p in process_list]
    node_colors = [
        get_process_color(detect_biodym_process_type(p.ID, dsm_params=dsm_params, fomp_params=fomp_params))
        for p in process_list
    ]

    sources, targets, values, hover = [], [], [], []
    for flow in mfa_system_results.FlowDict.values():
        value = float(flow.Values[year_idx, elem_idx])
        if value <= min_flow:
            continue
        if flow.P_Start not in process_id_map or flow.P_End not in process_id_map:
            continue
        sources.append(process_id_map[flow.P_Start])
        targets.append(process_id_map[flow.P_End])
        values.append(round(value, 6))
        hover.append(getattr(flow, "DescriptiveName", flow.Name))

    fig = go.Figure(go.Sankey(
        node=dict(label=node_labels, color=node_colors,
                  pad=sankey_config.NODE_SPACING,
                  thickness=sankey_config.NODE_THICKNESS),
        link=dict(source=sources, target=targets, value=values,
                  customdata=hover,
                  hovertemplate="%{customdata}<br />%{value:.1f} Mg<extra></extra>",
                  color=[link_color] * len(sources)),
    ))

    plot_title = title or f"BioDYM Sankey - {element} ({year})"
    fig.update_layout(title_text=plot_title, font_size=12, height=600)

    pathlib.Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    fig.write_html(str(filepath), include_plotlyjs="cdn")
    print(f"Sankey HTML exported: {filepath}  ({len(sources)} flows)")
    return fig


def export_sankey_csv(
    mfa_system_results,
    year,
    element,
    filepath,
    min_flow=0.0,
):
    """Export Sankey flows as a 3-column CSV compatible with e!Sankey (ifu Hamburg).

    Format (one flow per row, header included):
        Source,Target,Value
        Process A,Process B,1200.5

    Paste or import into e!Sankey via Data → Import.

    Args:
        mfa_system_results: Solved MFAsystem object.
        year (int): Calendar year to export (e.g., 2030).
        element (str): Element name (e.g., "material", "DM", "CC").
        filepath (str | Path): Output path, should end in .csv.
        min_flow (float): Minimum absolute flow value to include.

    Returns:
        list[dict]: The rows written (Source, Target, Value).
    """
    import csv
    import pathlib

    filepath = pathlib.Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    elements = mfa_system_results.Elements
    if element not in elements:
        raise ValueError(f"Element '{element}' not in system elements: {elements}.")
    time_vector = list(mfa_system_results.Time_V)
    if year not in time_vector:
        raise ValueError(f"Year {year} not in model time vector {time_vector}")
    t_idx = time_vector.index(year)
    el_idx = elements.index(element)
    name_map = {p.ID: p.Name for p in mfa_system_results.ProcessList}

    rows = []
    for flow in mfa_system_results.FlowDict.values():
        val = float(flow.Values[t_idx, el_idx])
        if abs(val) < min_flow:
            continue
        rows.append(
            {
                "Source": name_map.get(flow.P_Start, str(flow.P_Start)),
                "Target": name_map.get(flow.P_End, str(flow.P_End)),
                "Value": round(val, 4),
            }
        )

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["Source", "Target", "Value"])
        writer.writeheader()
        writer.writerows(rows)

    print(f"  ✅ e!Sankey CSV:     {filepath}  ({len(rows)} flows)")
    return rows


def export_sankey_sankeymatic(
    mfa_system_results,
    year,
    element,
    filepath,
    min_flow=0.0,
):
    """Export Sankey flows in SankeyMATIC plain-text format.

    Paste the file contents directly into sankeymatic.com.

    Format (one flow per line):
        Source [Value] Target
        Wheat Field [2000.0] Processing

    Multi-word node names work without modification (spaces are valid in SankeyMATIC).

    Args:
        mfa_system_results: Solved MFAsystem object.
        year (int): Calendar year to export.
        element (str): Element name (e.g., "material", "DM", "CC").
        filepath (str | Path): Output path, should end in .txt.
        min_flow (float): Minimum absolute flow value to include.

    Returns:
        str: The full text content written to the file.
    """
    import pathlib

    filepath = pathlib.Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    elements = mfa_system_results.Elements
    if element not in elements:
        raise ValueError(f"Element '{element}' not in system elements: {elements}.")
    time_vector = list(mfa_system_results.Time_V)
    if year not in time_vector:
        raise ValueError(f"Year {year} not in model time vector {time_vector}")
    t_idx = time_vector.index(year)
    el_idx = elements.index(element)
    name_map = {p.ID: p.Name for p in mfa_system_results.ProcessList}

    lines = [
        f"// BioDYM Sankey Export — {element} — {year}",
        "// Paste into sankeymatic.com",
        "",
    ]
    count = 0
    for flow in mfa_system_results.FlowDict.values():
        val = float(flow.Values[t_idx, el_idx])
        if abs(val) < min_flow:
            continue
        source = name_map.get(flow.P_Start, str(flow.P_Start))
        target = name_map.get(flow.P_End, str(flow.P_End))
        lines.append(f"{source} [{round(val, 2)}] {target}")
        count += 1

    content = "\n".join(lines)
    filepath.write_text(content, encoding="utf-8")

    print(f"  ✅ SankeyMATIC TXT:  {filepath}  ({count} flows)")
    return content


def export_mfa_diagram_xlsx(
    mfa_system_results,
    year,
    element,
    filepath,
    dsm_params=None,
    fomp_params=None,
    lfg_params=None,
    process_logic_map=None,
    node_positions=None,
    title="BioDYM MFA System",
    min_flow=0.0,
    stock_unit="Mg",
    flow_unit="Mg/yr",
):
    """Export MFA system to Structural Collective MFA Tool Excel format.

    Writes a .xlsx file compatible with https://mfa.structuralcollective.nl/editor
    with six sheets: Nodes, Flows, Groups, Settings, Node Colors, Node Positions.

    Parameters
    ----------
    mfa_system_results : odym.MFAsystem
        Solved MFA system with calculated flow and stock values.
    year : int
        Calendar year to export (single snapshot).
    element : str
        Element name to export (e.g. "material", "DM", "TC").
    filepath : str or Path
        Output path (should end in .xlsx).
    dsm_params : dict, optional
        DSM parameters dict — used to identify DSM processes for grouping/coloring.
    fomp_params : dict, optional
        FOMP parameters dict — used to identify FOMP processes.
    lfg_params : dict, optional
        LFG parameters dict — used to identify LFG processes.
    process_logic_map : dict, optional
        Maps process IDs to logic type strings (e.g. "DSM", "FOMP", "Input").
        When provided, used for group assignment instead of param dicts.
    node_positions : dict, optional
        Maps process names to (x, y) tuples for manual layout.
        Example: {"Atmosphere": (100, 200), "Field": (400, 200)}.
    title : str, optional
        Diagram title written to Settings sheet. Default "BioDYM MFA System".
    min_flow : float, optional
        Minimum absolute flow value to include. Default 0.0 (all non-zero).
    stock_unit : str, optional
        Unit label for stock values. Default "Mg".
    flow_unit : str, optional
        Unit label for flow values. Default "Mg/yr".

    Returns
    -------
    dict
        Summary with counts: {"nodes": n, "flows": n, "groups": n}.
    """
    import pathlib
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if dsm_params is None:
        dsm_params = {}
    if fomp_params is None:
        fomp_params = {}
    if lfg_params is None:
        lfg_params = {}
    if node_positions is None:
        node_positions = {}

    # --- Resolve year and element indices ---
    time_vector = list(mfa_system_results.Time_V)
    if year not in time_vector:
        raise ValueError(f"Year {year} not in model time range {time_vector[0]}–{time_vector[-1]}.")
    t_idx = time_vector.index(year)

    elements = mfa_system_results.Elements
    if element not in elements:
        raise ValueError(f"Element '{element}' not in system elements: {elements}.")
    el_idx = elements.index(element)

    name_map = {p.ID: p.Name for p in mfa_system_results.ProcessList}

    # --- Determine process type for each process ---
    def _proc_type(proc_id):
        if process_logic_map is not None:
            logic = process_logic_map.get(proc_id)
            if logic:
                return str(logic).strip()
        return detect_biodym_process_type(
            proc_id, dsm_params=dsm_params, fomp_params=fomp_params
        )

    # --- Collect nodes with stock values ---
    node_rows = []
    for p in mfa_system_results.ProcessList:
        stock_key = f"S_{p.ID}"
        stock_val = 0.0
        if stock_key in mfa_system_results.StockDict:
            sv = mfa_system_results.StockDict[stock_key].Values
            if sv is not None and sv.shape[0] > t_idx:
                stock_val = float(sv[t_idx, el_idx])
        node_rows.append({
            "node_id": p.Name,
            "title": p.Name,
            "stock": round(stock_val, 4),
            "stock_unit": stock_unit,
            "proc_id": p.ID,
        })

    # --- Collect flows ---
    flow_rows = []
    for flow in mfa_system_results.FlowDict.values():
        val = float(flow.Values[t_idx, el_idx])
        if abs(val) < min_flow:
            continue
        source = name_map.get(flow.P_Start, str(flow.P_Start))
        target = name_map.get(flow.P_End, str(flow.P_End))
        flow_rows.append({
            "source": source,
            "target": target,
            "value": round(val, 4),
            "flow_unit": flow_unit,
        })

    # --- Build groups: one row per (group_id, node) ---
    # Group logic: Input/boundary → "Boundary"; DSM/FOMP/LFG → own group; rest → "Process"
    _type_to_group = {
        "Input": "Boundary",
        "Output": "Boundary",
        "boundary": "Boundary",
        "input": "Boundary",
        "DSM": "DSM",
        "dsm": "DSM",
        "FOMP": "FOMP",
        "fomp": "FOMP",
        "LFG": "LFG",
        "lfg": "LFG",
        "Splitter": "Process",
        "regular": "Process",
    }
    group_rows = []
    group_name_map = {}
    for nd in node_rows:
        ptype = _proc_type(nd["proc_id"])
        grp_id = _type_to_group.get(ptype, "Process")
        group_rows.append({"group_id": grp_id, "group_name": grp_id, "node_id": nd["node_id"]})
        group_name_map[grp_id] = grp_id

    # --- Node colors from BioDYM theme ---
    color_rows = []
    for nd in node_rows:
        ptype = _proc_type(nd["proc_id"])
        color = get_process_color(ptype)
        color_rows.append({"node_title_or_id": nd["node_id"], "color": color})

    # --- Node positions ---
    pos_rows = []
    for nd in node_rows:
        pos = node_positions.get(nd["node_id"])
        if pos is not None:
            pos_rows.append({"node_id": nd["node_id"], "x": pos[0], "y": pos[1]})

    # --- Build workbook ---
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font = Font(bold=True)
    desc_font = Font(italic=True, color="888888")

    def _add_sheet(name, columns, desc_row, data_rows):
        ws = wb.create_sheet(name)
        ws.append(columns)
        ws.append(desc_row)
        for row in data_rows:
            ws.append([row.get(c, "") for c in columns])
        for col_idx in range(1, len(columns) + 1):
            ws.cell(1, col_idx).font = header_font
            ws.cell(2, col_idx).font = desc_font
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        return ws

    _add_sheet(
        "Nodes",
        ["node_id", "title", "stock", "stock_unit"],
        ["Unique node identifier", "Display name", "Stock value", "Unit for stock"],
        [{"node_id": r["node_id"], "title": r["title"],
          "stock": r["stock"], "stock_unit": r["stock_unit"]} for r in node_rows],
    )

    _add_sheet(
        "Flows",
        ["source", "target", "value", "flow_unit"],
        ["Source node ID", "Target node ID", "Flow value", "Unit for flow"],
        flow_rows,
    )

    _add_sheet(
        "Groups",
        ["group_id", "group_name", "node_id"],
        ["Unique group identifier", "Display name", "Member node ID"],
        group_rows,
    )

    # Settings sheet — key-value with defaults matching reference file
    ws_set = wb.create_sheet("Settings")
    ws_set.append(["Setting", "Value"])
    ws_set.cell(1, 1).font = header_font
    ws_set.cell(1, 2).font = header_font
    ws_set.column_dimensions["A"].width = 30
    ws_set.column_dimensions["B"].width = 30
    settings = [
        ("title", title),
        ("showFlowValues", True),
        ("showFlowUnits", True),
        ("showNodeValues", False),
        ("showStocks", True),
        ("showStockUnits", True),
        ("showInOutDetails", False),
        ("nodeValueNewLine", True),
        ("nodeLabelBackground", True),
        ("nodeLabelBackgroundOpacity", 0.85),
        ("nodePadding", 10),
        ("linkOpacity", 0.5),
        ("linkVisibility", 1),
        ("linkColorStyle", "solid"),
        ("colorMode", "origin"),
        ("clusterFlows", False),
        ("showGroupNames", True),
        ("showGroupIcons", False),
        ("showGroupTotals", True),
        ("canvasWidth", 1600),
        ("canvasHeight", 900),
        ("fontFamily", FONT_FAMILY),
        ("fontSize", 14),
        ("backgroundColor", "#f3f3f3"),
        ("textColor", "#000000"),
        ("valueScale", "none"),
        ("smallLinkThreshold", 1.5),
        ("smallLinkDotted", True),
        ("timeSeriesEnabled", False),
        ("nodeWidth", 15),
    ]
    for key, val in settings:
        ws_set.append([key, val])

    _add_sheet(
        "Node Colors",
        ["node_title_or_id", "color"],
        ["Node title or ID", "Hex color (e.g. #ff0000)"],
        color_rows,
    )

    _add_sheet(
        "Node Positions",
        ["node_id", "x", "y"],
        ["Node ID", "X position", "Y position"],
        pos_rows,
    )

    # --- Write file ---
    filepath = pathlib.Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)

    n_nodes = len(node_rows)
    n_flows = len(flow_rows)
    n_groups = len({r["group_id"] for r in group_rows})
    print(
        f"  Structural Collective MFA Tool xlsx exported: {filepath}\n"
        f"  {n_nodes} nodes, {n_flows} flows, {n_groups} groups  |  "
        f"year={year}, element={element}"
    )
    return {"nodes": n_nodes, "flows": n_flows, "groups": n_groups}
