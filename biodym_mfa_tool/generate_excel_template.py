# -*- coding: utf-8 -*-
"""
Excel Template Generator for BioDYM MFA Tool

This script generates a comprehensive Excel template with all required sheets
for the BioDYM MFA analysis tool.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import os

def create_excel_template(output_path='data/01_input/BioDYM_MFA_Input_Template.xlsx'):
    """
    Create a comprehensive Excel template for BioDYM MFA analysis.
    
    Args:
        output_path (str): Path where the Excel file should be saved.
    """
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ==============================================================================
    # SHEET 1: README
    # ==============================================================================
    ws_readme = wb.create_sheet("0_ReadMe")
    
    readme_data = [
        ["BioDYM MFA Tool - Excel Input Template"],
        [""],
        ["📋 SHEET OVERVIEW:"],
        ["   Configuration: Model settings and parameters"],
        ["   1_1_Definition_Flows: Define all material flows"],
        ["   1_2_Data_Flows: Time series data for flows"],
        ["   2_1_Definition_Processes: Define all processes"],
        ["   2_3_Process_TCs: Transfer coefficients for processes"],
        ["   2_4_Initial_Stock: Initial stock values"],
        ["   2_5_dynamic_tcs: Dynamic transfer coefficients"],
        ["   3_1_Definition_DSM: Dynamic Stock Model parameters"],
        ["   3_2_Definition_FOMP: First-Order Mineralization Process parameters"],
        ["   4_1_Uncertainty_Parameters: Monte Carlo uncertainty definitions"],
        [""],
        ["🎯 INSTRUCTIONS:"],
        ["1. Start with the Configuration sheet to set up your model"],
        ["2. Define your processes in 2_1_Definition_Processes"],
        ["3. Define your flows in 1_1_Definition_Flows"],
        ["4. Add time series data in 1_2_Data_Flows"],
        ["5. Configure DSM/FOMP parameters if needed"],
        ["6. Set up uncertainty parameters for Monte Carlo analysis"],
        [""],
        ["⚠️  IMPORTANT NOTES:"],
        ["- All sheets with green headers are required"],
        ["- Yellow cells are editable, gray cells are calculated"],
        ["- Use the data validation dropdowns where available"],
        ["- Save your work frequently"],
        [""],
        ["📞 SUPPORT:"],
        ["For questions or issues, refer to the BioDYM documentation."]
    ]
    
    for row in readme_data:
        ws_readme.append(row)
    
    # Style the README
    ws_readme['A1'].font = Font(bold=True, size=14)
    ws_readme['A3'].font = Font(bold=True, size=12)
    ws_readme['A15'].font = Font(bold=True, size=12)
    ws_readme['A22'].font = Font(bold=True, size=12)
    ws_readme['A27'].font = Font(bold=True, size=12)
    
    # ==============================================================================
    # SHEET 2: CONFIGURATION (NEW)
    # ==============================================================================
    ws_config = wb.create_sheet("Configuration")
    
    # Configuration data
    config_data = [
        ["BioDYM MFA Tool - Configuration Settings"],
        [""],
        ["📁 FILE PATHS"],
        ["Input File Path", "data/01_input/your_data_file.xlsx"],
        ["Output File Path", "data/02_output/results.xlsx"],
        [""],
        ["📊 MODEL SCOPE"],
        ["Start Year", 2025],
        ["End Year", 2050],
        ["Elements (comma-separated)", "material,WC,DM,CC"],
        [""],
        ["🎲 CALCULATION OPTIONS"],
        ["Run Monte Carlo Simulation", "No"],
        ["Monte Carlo Iterations", 100],
        ["Run DSM Calculation", "Yes"],
        ["Run FOMP Calculation", "Yes"],
        [""],
        ["📈 ANALYSIS SETTINGS"],
        ["Minimum Flow Threshold (Mg)", 0.1],
        ["Show Zero Flows in Plots", "No"],
        ["Export Format", "Excel"],
        [""],
        ["🎨 VISUALIZATION SETTINGS"],
        ["Default Plot Style", "Line"],
        ["Color Scheme", "Default"],
        ["Export Plots as Images", "Yes"],
        ["Dashboard Layout", "Grid"],
        [""],
        ["⚠️  VALIDATION SETTINGS"],
        ["Mass Balance Tolerance", 0.001],
        ["Data Validation Level", "Strict"],
        ["Auto-save Results", "Yes"]
    ]
    
    for row in config_data:
        ws_config.append(row)
    
    # Style the configuration sheet
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    section_fill = PatternFill(start_color="D0D8E4", end_color="D0D8E4", fill_type="solid")
    
    for row in range(1, len(config_data) + 1):
        cell = ws_config[f'A{row}']
        if row in [1, 3, 7, 12, 17, 22, 27]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        elif row in [4, 5, 8, 9, 10, 13, 14, 15, 16, 18, 19, 20, 23, 24, 25, 26, 28, 29, 30]:
            cell.fill = section_fill
    
    # Add data validation for dropdowns
    # Monte Carlo options
    mc_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws_config.add_data_validation(mc_validation)
    mc_validation.add('B13')  # Run Monte Carlo Simulation
    
    # Yes/No options
    yesno_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws_config.add_data_validation(yesno_validation)
    yesno_validation.add('B15')  # Run DSM Calculation
    yesno_validation.add('B16')  # Run FOMP Calculation
    yesno_validation.add('B20')  # Show Zero Flows in Plots
    yesno_validation.add('B22')  # Export Plots as Images
    yesno_validation.add('B29')  # Auto-save Results
    
    # Export format options
    export_validation = DataValidation(type="list", formula1='"Excel,CSV,JSON"', allow_blank=True)
    ws_config.add_data_validation(export_validation)
    export_validation.add('B21')  # Export Format
    
    # Plot style options
    plot_validation = DataValidation(type="list", formula1='"Line,Bar,Scatter"', allow_blank=True)
    ws_config.add_data_validation(plot_validation)
    plot_validation.add('B24')  # Default Plot Style
    
    # Color scheme options
    color_validation = DataValidation(type="list", formula1='"Default,Colorblind,High Contrast"', allow_blank=True)
    ws_config.add_data_validation(color_validation)
    color_validation.add('B25')  # Color Scheme
    
    # Dashboard layout options
    layout_validation = DataValidation(type="list", formula1='"Grid,Flow,Compact"', allow_blank=True)
    ws_config.add_data_validation(layout_validation)
    layout_validation.add('B27')  # Dashboard Layout
    
    # Data validation level options
    validation_level = DataValidation(type="list", formula1='"Strict,Moderate,Relaxed"', allow_blank=True)
    ws_config.add_data_validation(validation_level)
    validation_level.add('B30')  # Data Validation Level
    
    # ==============================================================================
    # SHEET 3: SCENARIO SETTINGS (NEW)
    # ==============================================================================
    ws_scenarios = wb.create_sheet("Scenario_Settings")
    
    scenario_data = [
        ["BioDYM MFA Tool - Scenario Settings"],
        [""],
        ["📋 CURRENT SCENARIO"],
        ["Scenario Name", "baseline"],
        ["Scenario Description", "Baseline scenario with current parameters"],
        ["Created Date", "2024-01-01"],
        ["Author", ""],
        [""],
        ["🔄 SCENARIO COMPARISON"],
        ["Compare Scenarios", "No"],
        ["Scenario 1", "baseline"],
        ["Scenario 2", "alternative"],
        ["Scenario 3", ""],
        [""],
        ["📊 COMPARISON METRICS"],
        ["Include Mass Balance", "Yes"],
        ["Include Stock Evolution", "Yes"],
        ["Include Flow Analysis", "Yes"],
        ["Include Efficiency Metrics", "Yes"],
        ["Include Monte Carlo Results", "Yes"],
        [""],
        ["🎯 SCENARIO PARAMETERS"],
        ["Parameter", "Baseline Value", "Alternative 1", "Alternative 2", "Description"],
        ["Start Year", 2025, 2025, 2025, "Analysis start year"],
        ["End Year", 2050, 2060, 2050, "Analysis end year"],
        ["MC Iterations", 100, 200, 50, "Monte Carlo iterations"],
        ["Recycling Rate", 0.8, 0.9, 0.7, "System recycling rate"],
        ["Efficiency Target", 0.85, 0.95, 0.75, "Target efficiency"]
    ]
    
    for row in scenario_data:
        ws_scenarios.append(row)
    
    # Style the scenario sheet
    for row in range(1, len(scenario_data) + 1):
        cell = ws_scenarios[f'A{row}']
        if row in [1, 3, 9, 15, 21]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        elif row in [4, 5, 6, 7, 10, 11, 12, 13, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26]:
            cell.fill = section_fill
    
    # Add data validation for scenario sheet
    compare_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws_scenarios.add_data_validation(compare_validation)
    compare_validation.add('B10')  # Compare Scenarios
    
    for col in ['B', 'C', 'D']:
        for row in [16, 17, 18, 19, 20]:
            yesno_validation.add(f'{col}{row}')  # Include metrics
    
    # ==============================================================================
    # SHEET 4: PROCESS DEFINITIONS
    # ==============================================================================
    ws_processes = wb.create_sheet("2_1_Definition_Processes")
    
    # Process definition headers
    process_headers = [
        "ID", "Name(EN)", "Stock?", "Initial_Stock?", "Process_Type", "Description"
    ]
    
    # Example processes
    process_data = [
        ["P_0", "Atmosphere", "No", "No", "Environment", "Atmospheric compartment"],
        ["P_1", "Environment", "No", "No", "Environment", "Environmental compartment"],
        ["P_2", "Cultivation", "No", "No", "Production", "Agricultural cultivation"],
        ["P_3", "Harvest", "No", "No", "Production", "Crop harvesting"],
        ["P_4", "Grain Processing & Consumption", "No", "No", "Processing", "Grain processing and consumption"],
        ["P_5", "Straw d&C", "No", "No", "Processing", "Straw decomposition and consumption"],
        ["P_6", "Utilization in construction", "Yes", "Yes", "Utilization", "Construction material use"],
        ["P_7", "Incineration", "No", "No", "Treatment", "Waste incineration"],
        ["P_8", "Incorporation", "No", "No", "Treatment", "Soil incorporation"],
        ["P_9", "Animal bedding", "No", "No", "Utilization", "Animal bedding use"],
        ["P_10", "Lithosphere", "Yes", "Yes", "Environment", "Soil compartment"]
    ]
    
    ws_processes.append(process_headers)
    for row in process_data:
        ws_processes.append(row)
    
    # Style process sheet
    for cell in ws_processes[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Add data validation
    stock_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws_processes.add_data_validation(stock_validation)
    
    process_type_validation = DataValidation(type="list", formula1='"Environment,Production,Processing,Utilization,Treatment,DSM,FOMP"', allow_blank=True)
    ws_processes.add_data_validation(process_type_validation)
    
    for row in range(2, len(process_data) + 2):
        stock_validation.add(f'C{row}')  # Stock?
        stock_validation.add(f'D{row}')  # Initial_Stock?
        process_type_validation.add(f'E{row}')  # Process_Type
    
    # ==============================================================================
    # SHEET 5: FLOW DEFINITIONS
    # ==============================================================================
    ws_flows = wb.create_sheet("1_1_Definition_Flows")
    
    # Flow definition headers
    flow_headers = [
        "Flow_ID", "Name(EN)", "Process_ID_O", "Process_ID_I", "Flow_Type", "Description"
    ]
    
    # Example flows
    flow_data = [
        ["F_00_02", "Atmosphere to Cultivation", "P_0", "P_2", "Input", "Atmospheric input to cultivation"],
        ["F_01_02", "Environment to Cultivation", "P_1", "P_2", "Input", "Environmental input to cultivation"],
        ["F_02_03", "Cultivation to Harvest", "P_2", "P_3", "Production", "Crop production"],
        ["F_03_04", "Harvest to Grain Processing", "P_3", "P_4", "Product", "Grain to processing"],
        ["F_03_05", "Harvest to Straw", "P_3", "P_5", "Product", "Straw from harvest"],
        ["F_04_00", "Grain Processing to Atmosphere", "P_4", "P_0", "Emission", "CO2 emission"],
        ["F_04_01", "Grain Processing to Environment", "P_4", "P_1", "Emission", "Waste emission"],
        ["F_05_06", "Straw to Construction", "P_5", "P_6", "Product", "Straw to construction"],
        ["F_06_07", "Construction to Incineration", "P_6", "P_7", "Waste", "Construction waste"],
        ["F_07_00", "Incineration to Atmosphere", "P_7", "P_0", "Emission", "CO2 from incineration"],
        ["F_07_01", "Incineration to Environment", "P_7", "P_1", "Emission", "Ash emission"],
        ["F_05_08", "Straw to Incorporation", "P_5", "P_8", "Product", "Straw to soil"],
        ["F_08_10", "Incorporation to Lithosphere", "P_8", "P_10", "Transfer", "Soil incorporation"],
        ["F_05_09", "Straw to Animal Bedding", "P_5", "P_9", "Product", "Straw to bedding"],
        ["F_09_00", "Animal Bedding to Atmosphere", "P_9", "P_0", "Emission", "CO2 from bedding"],
        ["F_09_01", "Animal Bedding to Environment", "P_9", "P_1", "Emission", "Waste from bedding"]
    ]
    
    ws_flows.append(flow_headers)
    for row in flow_data:
        ws_flows.append(row)
    
    # Style flow sheet
    for cell in ws_flows[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Add data validation for flow types
    flow_type_validation = DataValidation(type="list", formula1='"Input,Output,Product,Waste,Emission,Transfer"', allow_blank=True)
    ws_flows.add_data_validation(flow_type_validation)
    
    for row in range(2, len(flow_data) + 2):
        flow_type_validation.add(f'E{row}')  # Flow_Type
    
    # ==============================================================================
    # SHEET 6: FLOW DATA
    # ==============================================================================
    ws_flow_data = wb.create_sheet("1_2_Data_Flows")
    
    # Flow data headers
    flow_data_headers = [
        "Flow_ID", "Year_Flow", "Flow_Py", "Flow_WC", "Flow_DM", "Flow_CC"
    ]
    
    # Example flow data (first few years)
    flow_data_values = []
    for flow_id in ["F_00_02", "F_01_02", "F_02_03", "F_03_04", "F_03_05"]:
        for year in range(2025, 2031):  # 6 years of data
            flow_data_values.append([
                flow_id, year, 
                100 + (year - 2025) * 10,  # material
                50 + (year - 2025) * 5,   # WC
                30 + (year - 2025) * 3,   # DM
                20 + (year - 2025) * 2    # CC
            ])
    
    ws_flow_data.append(flow_data_headers)
    for row in flow_data_values:
        ws_flow_data.append(row)
    
    # Style flow data sheet
    for cell in ws_flow_data[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # ==============================================================================
    # SHEET 7: TRANSFER COEFFICIENTS
    # ==============================================================================
    ws_tcs = wb.create_sheet("2_3_Process_TCs")
    
    # TC headers
    tc_headers = [
        "Process_ID", "Flow_ID_O", "Flow_ID_I", "TC_Value", "Description"
    ]
    
    # Example TCs
    tc_data = [
        ["P_3", "F_02_03", "F_03_04", 0.5, "50% of harvest to grain"],
        ["P_3", "F_02_03", "F_03_05", 0.5, "50% of harvest to straw"],
        ["P_4", "F_03_04", "F_04_00", 0.8, "80% of grain to CO2"],
        ["P_4", "F_03_04", "F_04_01", 0.2, "20% of grain to waste"],
        ["P_5", "F_03_05", "F_05_06", 0.4, "40% of straw to construction"],
        ["P_5", "F_03_05", "F_05_08", 0.3, "30% of straw to soil"],
        ["P_5", "F_03_05", "F_05_09", 0.3, "30% of straw to bedding"]
    ]
    
    ws_tcs.append(tc_headers)
    for row in tc_data:
        ws_tcs.append(row)
    
    # Style TC sheet
    for cell in ws_tcs[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # ==============================================================================
    # SHEET 8: INITIAL STOCKS
    # ==============================================================================
    ws_stocks = wb.create_sheet("2_4_Initial_Stock")
    
    # Stock headers
    stock_headers = [
        "Process_ID", "Initial_Stock_material", "Initial_Stock_WC", "Initial_Stock_DM", "Initial_Stock_CC", "Description"
    ]
    
    # Example stocks
    stock_data = [
        ["P_6", 0, 0, 0, 0, "Construction stock (starts empty)"],
        ["P_10", 0, 0, 0, 0, "Soil stock (starts empty)"]
    ]
    
    ws_stocks.append(stock_headers)
    for row in stock_data:
        ws_stocks.append(row)
    
    # Style stock sheet
    for cell in ws_stocks[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # ==============================================================================
    # SHEET 9: DYNAMIC TRANSFER COEFFICIENTS
    # ==============================================================================
    ws_dynamic_tcs = wb.create_sheet("2_5_dynamic_tcs")
    
    # Dynamic TC headers
    dynamic_tc_headers = [
        "TC_ID", "Year", "Value", "Description"
    ]
    
    # Example dynamic TCs
    dynamic_tc_data = [
        ["TC_1", 2025, 0.5, "Initial TC value"],
        ["TC_1", 2030, 0.6, "Increased efficiency"],
        ["TC_1", 2035, 0.7, "Further improvement"],
        ["TC_1", 2040, 0.8, "Target efficiency"],
        ["TC_1", 2045, 0.85, "High efficiency"],
        ["TC_1", 2050, 0.9, "Maximum efficiency"]
    ]
    
    ws_dynamic_tcs.append(dynamic_tc_headers)
    for row in dynamic_tc_data:
        ws_dynamic_tcs.append(row)
    
    # Style dynamic TC sheet
    for cell in ws_dynamic_tcs[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # ==============================================================================
    # SHEET 10: DSM DEFINITION
    # ==============================================================================
    ws_dsm = wb.create_sheet("3_1_Definition_DSM")
    
    # DSM headers
    dsm_headers = [
        "Process_ID", "DSM_Type", "Decay_Rate", "Description"
    ]
    
    # Example DSM (empty for now)
    dsm_data = [
        # ["P_6", "Exponential", 0.1, "Construction material decay"]
    ]
    
    ws_dsm.append(dsm_headers)
    for row in dsm_data:
        ws_dsm.append(row)
    
    # Style DSM sheet
    for cell in ws_dsm[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Add data validation for DSM types
    dsm_type_validation = DataValidation(type="list", formula1='"Exponential,Linear,Weibull"', allow_blank=True)
    ws_dsm.add_data_validation(dsm_type_validation)
    
    # ==============================================================================
    # SHEET 11: FOMP DEFINITION
    # ==============================================================================
    ws_fomp = wb.create_sheet("3_2_Definition_FOMP")
    
    # FOMP headers
    fomp_headers = [
        "Process_ID", "Outflow_ID", "f", "k1", "k2", "Description"
    ]
    
    # Example FOMP
    fomp_data = [
        ["P_10", "F_08_10", 0.236, 0.025, 0.0351, "Soil organic matter decomposition"]
    ]
    
    ws_fomp.append(fomp_headers)
    for row in fomp_data:
        ws_fomp.append(row)
    
    # Style FOMP sheet
    for cell in ws_fomp[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # ==============================================================================
    # SHEET 12: UNCERTAINTY PARAMETERS
    # ==============================================================================
    ws_uncertainty = wb.create_sheet("4_1_Uncertainty_Parameters")
    
    # Uncertainty headers
    uncertainty_headers = [
        "Parameter_ID", "Parameter_Name", "Distribution", "Mean", "Std_Dev", "Min", "Max", "Description"
    ]
    
    # Example uncertainty parameters
    uncertainty_data = [
        ["TC_1", "Transfer Coefficient 1", "Normal", 0.5, 0.1, 0.3, 0.7, "Uncertainty in TC1"],
        ["TC_2", "Transfer Coefficient 2", "Normal", 0.3, 0.05, 0.2, 0.4, "Uncertainty in TC2"],
        ["F_02_03", "Cultivation to Harvest Flow", "Normal", 200, 20, 160, 240, "Uncertainty in harvest flow"],
        ["k1", "FOMP k1 parameter", "Normal", 0.025, 0.005, 0.015, 0.035, "Uncertainty in FOMP k1"]
    ]
    
    ws_uncertainty.append(uncertainty_headers)
    for row in uncertainty_data:
        ws_uncertainty.append(row)
    
    # Style uncertainty sheet
    for cell in ws_uncertainty[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Add data validation for distributions
    dist_validation = DataValidation(type="list", formula1='"Normal,Uniform,LogNormal,Triangular"', allow_blank=True)
    ws_uncertainty.add_data_validation(dist_validation)
    
    for row in range(2, len(uncertainty_data) + 2):
        dist_validation.add(f'C{row}')  # Distribution
    
    # ==============================================================================
    # SAVE THE WORKBOOK
    # ==============================================================================
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel template created successfully: {output_path}")
    print(f"Template includes {len(wb.sheetnames)} sheets:")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")
    
    return output_path

if __name__ == "__main__":
    # Create the template
    template_path = create_excel_template()
    print(f"\nTemplate ready for use!")
    print(f"Next steps:")
    print(f"   1. Open {template_path}")
    print(f"   2. Configure your model in the 'Configuration' sheet")
    print(f"   3. Define your processes and flows")
    print(f"   4. Add your time series data")
    print(f"   5. Run your MFA analysis!") 