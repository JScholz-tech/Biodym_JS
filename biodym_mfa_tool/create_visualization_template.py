#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Template Generator for BioDYM Visualization Configuration

This script creates a comprehensive Excel template with:
1. Enhanced process definition sheet with visualization columns
2. New visualization configuration sheet
3. Code lists for dropdown validation
4. Enhanced flow definition sheet
5. Data validation rules and formatting

Run this script to generate the Excel template file.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def create_visualization_template():
    """Create the comprehensive visualization configuration Excel template."""
    
    # Create output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"BioDYM_Visualization_Template_{timestamp}.xlsx"
    
    print(f"🔧 Creating BioDYM Visualization Template: {output_file}")
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # ==============================================================================
        # SHEET 1: Enhanced Process Definition (2_1_Definition_Processes_Enhanced)
        # ==============================================================================
        print("📋 Creating enhanced process definition sheet...")
        
        enhanced_processes_data = {
            'ID': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            'Name(EN)': [
                'Biomass Input', 'Primary Processing', 'Storage Tank', 
                'Secondary Processing', 'Quality Control', 'Packaging',
                'Distribution', 'Waste Treatment', 'Recycling', 'Final Output'
            ],
            'Stock?': ['No', 'No', 'Yes', 'No', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Initial_Stock?': ['No', 'No', 'Yes', 'No', 'No', 'No', 'No', 'No', 'No', 'No'],
            'Process_Category': ['Input', 'Transformation', 'Storage', 'Transformation', 
                               'Transformation', 'Transformation', 'Transformation', 
                               'Transformation', 'Transformation', 'Output'],
            'Process_Layer': [1, 2, 3, 2, 3, 4, 4, 3, 3, 5],
            'Cluster_ID': [1, 2, 2, 2, 3, 3, 4, 5, 5, 6],
            'Visual_Shape': ['Circle', 'Square', 'Diamond', 'Square', 'Square', 
                            'Square', 'Square', 'Hexagon', 'Hexagon', 'Circle'],
            'Visual_Color': ['Blue', 'Orange', 'Green', 'Orange', 'Orange', 
                            'Orange', 'Purple', 'Red', 'Red', 'Red'],
            'Show_Details': ['Always', 'On_Hover', 'Always', 'On_Hover', 'On_Click', 
                            'On_Click', 'On_Hover', 'Collapsed', 'Collapsed', 'Always'],
            'Layout_Priority': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            'Description': [
                'Raw biomass material input to the system',
                'Primary mechanical and chemical processing',
                'Intermediate storage with temperature control',
                'Secondary processing and refinement',
                'Quality testing and validation',
                'Final product packaging',
                'Distribution and logistics',
                'Waste treatment and disposal',
                'Material recycling and recovery',
                'Final product output from system'
            ]
        }
        
        enhanced_processes_df = pd.DataFrame(enhanced_processes_data)
        enhanced_processes_df.to_excel(writer, sheet_name='2_1_Definition_Processes_Enhanced', index=False)
        
        # ==============================================================================
        # SHEET 2: Visualization Configuration (5_1_Visualization_Configuration)
        # ==============================================================================
        print("⚙️ Creating visualization configuration sheet...")
        
        viz_config_data = {
            'Setting_Name': [
                'Default Layout Type',
                'Default Process Category Colors',
                'Default Visual Shapes',
                'Default Detail Level',
                'Enable Process Clustering',
                'Enable Flow Bundling',
                'Default Edge Routing',
                'Show Flow Labels',
                'Show Process Icons',
                'Enable Progressive Disclosure',
                'Default Color Scheme',
                'Layout Spacing',
                'Node Size',
                'Edge Thickness',
                'Animation Speed',
                'Export Quality',
                'Default View Mode',
                'Enable Zoom Controls',
                'Show Legend',
                'Enable Tooltips'
            ],
            'Setting_Value': [
                'Hierarchical',
                'Yes',
                'Yes',
                'On_Hover',
                'Yes',
                'Yes',
                'Orthogonal',
                'Yes',
                'Yes',
                'Yes',
                'Professional',
                'Medium',
                'Medium',
                'Medium',
                'Normal',
                'High',
                'Interactive',
                'Yes',
                'Yes',
                'Yes'
            ],
            'Setting_Description': [
                'Main layout strategy for Sankey diagram (Hierarchical, Circular, Force_Directed, Cluster_Based)',
                'Use predefined colors for process categories',
                'Use different shapes for different process types',
                'Default detail level for process information (Always, On_Hover, On_Click, Collapsed)',
                'Group related processes into visual clusters',
                'Bundle parallel flows to reduce visual clutter',
                'Edge routing style (Direct, Orthogonal, Curved, Bundled)',
                'Display flow labels on edges',
                'Show process type icons on nodes',
                'Show detailed information progressively',
                'Color scheme for the visualization (Professional, Vibrant, Monochrome, Custom)',
                'Spacing between layout elements (Tight, Medium, Loose)',
                'Size of process nodes (Small, Medium, Large)',
                'Thickness of flow edges (Thin, Medium, Thick)',
                'Speed of layout animations (Slow, Normal, Fast)',
                'Quality for exported images (Low, Medium, High, Vector)',
                'Default view mode (Interactive, Static, Presentation)',
                'Enable zoom and pan controls',
                'Display color and shape legend',
                'Show detailed information on hover'
            ],
            'Category': [
                'Layout',
                'Visual',
                'Visual',
                'Interaction',
                'Layout',
                'Layout',
                'Layout',
                'Visual',
                'Visual',
                'Interaction',
                'Visual',
                'Layout',
                'Visual',
                'Visual',
                'Interaction',
                'Export',
                'View',
                'Interaction',
                'Visual',
                'Interaction'
            ],
            'Is_Active': [
                'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes',
                'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'
            ],
            'Priority': [
                1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20
            ]
        }
        
        viz_config_df = pd.DataFrame(viz_config_data)
        viz_config_df.to_excel(writer, sheet_name='5_1_Visualization_Configuration', index=False)
        
        # ==============================================================================
        # SHEET 3: Code Lists (5_2_Visualization_Code_Lists)
        # ==============================================================================
        print("📚 Creating code lists sheet...")
        
        # Create a simplified, consistent code list structure
        code_lists_data = {
            'Code_List_Name': [
                'Process_Category', 'Process_Category', 'Process_Category', 'Process_Category', 'Process_Category',
                'Visual_Shape', 'Visual_Shape', 'Visual_Shape', 'Visual_Shape', 'Visual_Shape',
                'Visual_Color', 'Visual_Color', 'Visual_Color', 'Visual_Color', 'Visual_Color',
                'Layout_Type', 'Layout_Type', 'Layout_Type', 'Layout_Type', 'Layout_Type',
                'Detail_Level', 'Detail_Level', 'Detail_Level', 'Detail_Level', 'Detail_Level'
            ],
            'Code_Value': [
                'Input', 'Transformation', 'Storage', 'Output', 'Interface',
                'Circle', 'Square', 'Diamond', 'Hexagon', 'Triangle',
                'Blue', 'Orange', 'Green', 'Red', 'Purple',
                'Hierarchical', 'Circular', 'Force_Directed', 'Cluster_Based', 'Grid',
                'Always', 'On_Hover', 'On_Click', 'Collapsed', 'Default'
            ],
            'Description': [
                'Input processes (raw materials, external sources)',
                'Processing and transformation processes',
                'Storage and stock processes',
                'Output processes (final products, disposal)',
                'System boundary and interface processes',
                'Circular node shape',
                'Square node shape',
                'Diamond node shape',
                'Hexagonal node shape',
                'Triangular node shape',
                'Primary color for main processes',
                'Secondary color for transformation processes',
                'Success color for storage processes',
                'Warning color for output/waste processes',
                'Accent color for special processes',
                'Left-to-right hierarchical arrangement',
                'Circular arrangement around center',
                'Force-based dynamic positioning',
                'Grouped by process clusters',
                'Regular grid arrangement',
                'Always show full details',
                'Show details on mouse hover',
                'Show details on click',
                'Hide details by default',
                'Use system default setting'
            ],
            'Color_Code': [
                '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                '#000000', '#000000', '#000000', '#000000', '#000000',
                '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                '#000000', '#000000', '#000000', '#000000', '#000000',
                '#000000', '#000000', '#000000', '#000000', '#000000'
            ],
            'Icon_Symbol': [
                '🔵', '🟠', '🟢', '🔴', '🟣',
                '⭕', '⬜', '💎', '⬡', '🔺',
                '🔵', '🟠', '🟢', '🔴', '🟣',
                '➡️', '🔄', '⚡', '🗂️', '⬜',
                '👁️', '👁️', '👁️', '📁', '⚙️'
            ]
        }
        
        code_lists_df = pd.DataFrame(code_lists_data)
        code_lists_df.to_excel(writer, sheet_name='5_2_Visualization_Code_Lists', index=False)
        
        # ==============================================================================
        # SHEET 4: Enhanced Flow Definition (1_1_Definition_Flows_Enhanced)
        # ==============================================================================
        print("🔄 Creating enhanced flow definition sheet...")
        
        enhanced_flows_data = {
            'Flow_ID': ['F_01_02', 'F_02_03', 'F_03_04', 'F_04_05', 'F_05_06', 
                       'F_06_07', 'F_07_08', 'F_02_09', 'F_04_09', 'F_08_10'],
            'Name(EN)': [
                'Biomass to Processing', 'Processed to Storage', 'Storage to Secondary',
                'Secondary to Quality', 'Quality to Packaging', 'Packaging to Distribution',
                'Distribution to Output', 'Processing Waste', 'Secondary Waste', 'Waste to Treatment'
            ],
            'Process_ID_O': [1, 2, 3, 4, 5, 6, 7, 2, 4, 8],
            'Process_ID_I': [2, 3, 4, 5, 6, 7, 8, 9, 9, 10],
            'Flow_Category': ['Primary', 'Primary', 'Primary', 'Primary', 'Primary', 
                             'Primary', 'Primary', 'Waste', 'Waste', 'Waste'],
            'Flow_Importance': [1, 1, 1, 2, 2, 2, 2, 3, 3, 3],
            'Routing_Preference': ['Direct', 'Direct', 'Direct', 'Direct', 'Direct', 
                                  'Direct', 'Direct', 'Curved', 'Curved', 'Curved'],
            'Flow_Group': ['Group_1', 'Group_1', 'Group_1', 'Group_2', 'Group_2', 
                           'Group_2', 'Group_2', 'Group_3', 'Group_3', 'Group_3'],
            'Visual_Thickness': ['Thick', 'Medium', 'Medium', 'Medium', 'Medium', 
                                 'Medium', 'Medium', 'Thin', 'Thin', 'Thin'],
            'Show_Label': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'No', 'No', 'No'],
            'Description': [
                'Main biomass flow from input to primary processing',
                'Processed material flow to storage tank',
                'Material flow from storage to secondary processing',
                'Flow from secondary processing to quality control',
                'Quality-approved material to packaging',
                'Packaged product to distribution',
                'Distribution to final output',
                'Waste flow from primary processing',
                'Waste flow from secondary processing',
                'Combined waste to treatment'
            ]
        }
        
        enhanced_flows_df = pd.DataFrame(enhanced_flows_data)
        enhanced_flows_df.to_excel(writer, sheet_name='1_1_Definition_Flows_Enhanced', index=False)
        
        # ==============================================================================
        # SHEET 5: Layout Templates (5_3_Layout_Templates)
        # ==============================================================================
        print("📐 Creating layout templates sheet...")
        
        layout_templates_data = {
            'Template_Name': [
                'Standard Hierarchical',
                'Compact Cluster',
                'Presentation Mode',
                'Technical Detail',
                'Stakeholder View',
                'Process Focus',
                'Flow Focus',
                'Storage Focus',
                'Waste Focus',
                'Custom Template'
            ],
            'Layout_Type': [
                'Hierarchical',
                'Cluster_Based',
                'Hierarchical',
                'Force_Directed',
                'Hierarchical',
                'Hierarchical',
                'Circular',
                'Grid',
                'Hierarchical',
                'Custom'
            ],
            'Process_Category_Colors': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'],
            'Visual_Shapes': ['Yes', 'Yes', 'No', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'],
            'Detail_Level': ['On_Hover', 'Collapsed', 'Always', 'Always', 'On_Hover', 'On_Click', 'On_Hover', 'Always', 'On_Hover', 'Custom'],
            'Process_Clustering': ['No', 'Yes', 'No', 'Yes', 'No', 'No', 'Yes', 'Yes', 'No', 'Custom'],
            'Flow_Bundling': ['No', 'Yes', 'No', 'Yes', 'No', 'No', 'Yes', 'Yes', 'No', 'Custom'],
            'Edge_Routing': ['Orthogonal', 'Bundled', 'Direct', 'Curved', 'Orthogonal', 'Orthogonal', 'Curved', 'Orthogonal', 'Orthogonal', 'Custom'],
            'Show_Labels': ['Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Custom'],
            'Show_Icons': ['Yes', 'Yes', 'No', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'Custom'],
            'Progressive_Disclosure': ['Yes', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Custom'],
            'Description': [
                'Standard left-to-right hierarchical layout for general use',
                'Compact layout with process clustering for complex systems',
                'High-detail layout suitable for presentations and reports',
                'Technical layout with maximum detail and force-directed positioning',
                'Simplified layout for stakeholder communication',
                'Layout emphasizing process details and relationships',
                'Layout emphasizing flow patterns and circular arrangements',
                'Grid-based layout emphasizing storage and stock processes',
                'Layout emphasizing waste flows and treatment processes',
                'User-defined custom layout configuration'
            ],
            'Use_Case': [
                'General analysis and documentation',
                'Complex system analysis with many processes',
                'Presentations and stakeholder meetings',
                'Technical analysis and debugging',
                'High-level stakeholder communication',
                'Process optimization and analysis',
                'Flow pattern analysis and optimization',
                'Storage and inventory analysis',
                'Waste management and treatment analysis',
                'Specialized analysis requirements'
            ]
        }
        
        layout_templates_df = pd.DataFrame(layout_templates_data)
        layout_templates_df.to_excel(writer, sheet_name='5_3_Layout_Templates', index=False)
        
        # ==============================================================================
        # SHEET 6: Implementation Guide (5_4_Implementation_Guide)
        # ==============================================================================
        print("📖 Creating implementation guide sheet...")
        
        implementation_guide_data = {
            'Step': [
                '1. Process Categorization',
                '2. Layout Configuration',
                '3. Visual Settings',
                '4. Flow Configuration',
                '5. Template Selection',
                '6. Data Validation',
                '7. Testing',
                '8. Customization',
                '9. Documentation',
                '10. Maintenance'
            ],
            'Action': [
                'Categorize all processes in 2_1_Definition_Processes_Enhanced',
                'Configure layout settings in 5_1_Visualization_Configuration',
                'Set visual preferences for processes and flows',
                'Configure flow properties in 1_1_Definition_Flows_Enhanced',
                'Select appropriate layout template from 5_3_Layout_Templates',
                'Apply data validation rules to all sheets',
                'Test visualization with sample data',
                'Adjust settings based on specific requirements',
                'Document custom configurations and settings',
                'Regular review and update of visualization settings'
            ],
            'Details': [
                'Assign Process_Category, Process_Layer, and Cluster_ID to each process',
                'Set Default Layout Type, Enable Process Clustering, and other layout options',
                'Configure Visual_Shape, Visual_Color, and Show_Details for each process',
                'Set Flow_Category, Flow_Importance, and Routing_Preference for each flow',
                'Choose template based on use case and audience requirements',
                'Ensure all dropdown menus work correctly and data is consistent',
                'Verify that visualizations render correctly with your data',
                'Modify colors, shapes, and layouts to match your brand or preferences',
                'Keep record of any custom settings for future reference',
                'Update visualization settings as your system evolves'
            ],
            'Required_Sheets': [
                '2_1_Definition_Processes_Enhanced',
                '5_1_Visualization_Configuration',
                '2_1_Definition_Processes_Enhanced',
                '1_1_Definition_Flows_Enhanced',
                '5_3_Layout_Templates',
                'All sheets with data validation',
                'All sheets with sample data',
                'All configuration sheets',
                '5_4_Implementation_Guide',
                'All sheets'
            ],
            'Estimated_Time': [
                '30 minutes',
                '15 minutes',
                '20 minutes',
                '20 minutes',
                '10 minutes',
                '15 minutes',
                '30 minutes',
                '45 minutes',
                '20 minutes',
                '15 minutes'
            ]
        }
        
        implementation_guide_df = pd.DataFrame(implementation_guide_data)
        implementation_guide_df.to_excel(writer, sheet_name='5_4_Implementation_Guide', index=False)
    
    print(f"✅ Excel template created successfully: {output_file}")
    print("\n📋 Template contains the following sheets:")
    print("   1. 2_1_Definition_Processes_Enhanced - Enhanced process definitions")
    print("   2. 5_1_Visualization_Configuration - Main configuration settings")
    print("   3. 5_2_Visualization_Code_Lists - Dropdown options and validation")
    print("   4. 1_1_Definition_Flows_Enhanced - Enhanced flow definitions")
    print("   5. 5_3_Layout_Templates - Pre-configured layout templates")
    print("   6. 5_4_Implementation_Guide - Step-by-step implementation guide")
    
    return output_file

def add_data_validation_to_excel(filename):
    """Add data validation rules to the Excel file."""
    
    print(f"\n🔧 Adding data validation rules to {filename}...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # ==============================================================================
    # Add data validation to Process Definition sheet
    # ==============================================================================
    if '2_1_Definition_Processes_Enhanced' in wb.sheetnames:
        ws = wb['2_1_Definition_Processes_Enhanced']
        
        # Process Category validation (Column E)
        dv = DataValidation(type="list", formula1='"Input,Transformation,Storage,Output,Interface"', allow_blank=True)
        dv.add(f'E2:E{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Process Layer validation (Column F)
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
        dv.add(f'F2:F{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Cluster ID validation (Column G)
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
        dv.add(f'G2:G{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Visual Shape validation (Column H)
        dv = DataValidation(type="list", formula1='"Circle,Square,Diamond,Hexagon,Triangle"', allow_blank=True)
        dv.add(f'H2:H{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Visual Color validation (Column I)
        dv = DataValidation(type="list", formula1='"Blue,Orange,Green,Red,Purple,Gray"', allow_blank=True)
        dv.add(f'I2:I{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Show Details validation (Column J)
        dv = DataValidation(type="list", formula1='"Always,On_Hover,On_Click,Collapsed"', allow_blank=True)
        dv.add(f'J2:J{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Layout Priority validation (Column K)
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
        dv.add(f'K2:K{ws.max_row}')
        ws.add_data_validation(dv)
        
        print("✅ Added data validation to Process Definition sheet")
    
    # ==============================================================================
    # Add data validation to Flow Definition sheet
    # ==============================================================================
    if '1_1_Definition_Flows_Enhanced' in wb.sheetnames:
        ws = wb['1_1_Definition_Flows_Enhanced']
        
        # Flow Category validation (Column E)
        dv = DataValidation(type="list", formula1='"Primary,Secondary,Tertiary,Waste"', allow_blank=True)
        dv.add(f'E2:E{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Flow Importance validation (Column F)
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
        dv.add(f'F2:F{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Routing Preference validation (Column G)
        dv = DataValidation(type="list", formula1='"Direct,Orthogonal,Curved,Bundled"', allow_blank=True)
        dv.add(f'G2:G{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Visual Thickness validation (Column I)
        dv = DataValidation(type="list", formula1='"Thin,Medium,Thick"', allow_blank=True)
        dv.add(f'I2:I{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Show Label validation (Column J)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.add(f'J2:J{ws.max_row}')
        ws.add_data_validation(dv)
        
        print("✅ Added data validation to Flow Definition sheet")
    
    # ==============================================================================
    # Add data validation to Visualization Configuration sheet
    # ==============================================================================
    if '5_1_Visualization_Configuration' in wb.sheetnames:
        ws = wb['5_1_Visualization_Configuration']
        
        # Setting Value validation (Column B) - This will be dynamic based on setting name
        # For now, add basic validation
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.add(f'B2:B{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Category validation (Column D)
        dv = DataValidation(type="list", formula1='"Layout,Visual,Interaction,Export,View"', allow_blank=True)
        dv.add(f'D2:D{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Is Active validation (Column E)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.add(f'E2:E{ws.max_row}')
        ws.add_data_validation(dv)
        
        # Priority validation (Column F)
        dv = DataValidation(type="whole", operator="between", formula1="1", formula2="20", allow_blank=True)
        dv.add(f'F2:F{ws.max_row}')
        ws.add_data_validation(dv)
        
        print("✅ Added data validation to Visualization Configuration sheet")
    
    # Save the workbook with validation
    wb.save(filename)
    print(f"✅ Data validation rules added and saved to {filename}")

def format_excel_file(filename):
    """Apply professional formatting to the Excel file."""
    
    print(f"\n🎨 Applying professional formatting to {filename}...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting to all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = data_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Formatted sheet: {sheet_name}")
    
    # Save the formatted workbook
    wb.save(filename)
    print(f"✅ Professional formatting applied and saved to {filename}")

if __name__ == "__main__":
    print("🚀 BioDYM Visualization Configuration Template Generator")
    print("=" * 60)
    
    try:
        # Create the template
        template_file = create_visualization_template()
        
        # Add data validation
        add_data_validation_to_excel(template_file)
        
        # Apply formatting
        format_excel_file(template_file)
        
        print("\n🎉 Template generation completed successfully!")
        print(f"📁 Your template is ready: {template_file}")
        print("\n📋 Next steps:")
        print("   1. Review the template structure")
        print("   2. Customize the example data for your case study")
        print("   3. Apply the visualization settings you prefer")
        print("   4. Test with your BioDYM system")
        print("   5. Share feedback for further improvements")
        
    except Exception as e:
        print(f"❌ Error generating template: {e}")
        import traceback
        traceback.print_exc()
