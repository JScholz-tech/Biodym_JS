import pandas as pd
import openpyxl

# Load the Excel file to get current column names
file_path = 'data/01_input/251014_BioDYM_DataProtocoll_CS2_Wood.xlsx'

# Read validation sheet to get all column names
df_val = pd.read_excel(file_path, sheet_name='7_1_Comments_Validation', header=2)

print("COMPREHENSIVE COLUMN MAPPING ANALYSIS")
print("="*80)

# Get all current column names from validation sheet
current_columns = {}
for _, row in df_val.iterrows():
    if pd.notna(row['Name_sheet']) and pd.notna(row['Name_Column']):
        sheet = row['Name_sheet']
        col = row['Name_Column']
        if sheet not in current_columns:
            current_columns[sheet] = []
        current_columns[sheet].append(col)

print("CURRENT COLUMN NAMES BY SHEET:")
print("-" * 50)

# Also read actual Excel sheets to verify
wb = openpyxl.load_workbook(file_path)
actual_columns = {}

for sheet_name in wb.sheetnames:
    if sheet_name in current_columns:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            actual_columns[sheet_name] = list(df.columns)
            print(f"\n{sheet_name}:")
            print(f"  Validation sheet: {len(current_columns[sheet_name])} columns")
            print(f"  Actual sheet: {len(actual_columns[sheet_name])} columns")
            
            # Check for differences
            val_cols = set(current_columns[sheet_name])
            actual_cols = set(actual_columns[sheet_name])
            
            missing_in_actual = val_cols - actual_cols
            missing_in_validation = actual_cols - val_cols
            
            if missing_in_actual:
                print(f"  Missing in actual: {missing_in_actual}")
            if missing_in_validation:
                print(f"  Missing in validation: {missing_in_validation}")
                
        except Exception as e:
            print(f"\n{sheet_name}: Error reading - {e}")

print("\n\nCRITICAL COLUMN MAPPINGS NEEDED:")
print("="*80)

# Based on the data loader analysis, here are the critical columns that need mapping
critical_mappings = {
    "1_1_Definition_Flows": {
        "Flow_ID": "Flow_ID",  # No change needed
        "Flow_Name": "Flow_Name",  # No change needed
        "Flow_Output_Process_ID": "Flow_Output_Process_ID",  # No change needed
        "Input_Process_ID": "Input_Process_ID",  # No change needed
        "Flow_WC_ID ": "Flow_WC_ID",  # Remove trailing space
        "Complete?": "Complete?",  # No change needed
        "WC?": "WC?",  # No change needed
        "DM?": "DM?",  # No change needed
        "CC?": "CC?",  # No change needed
    },
    "1_2_Data_Flows": {
        "Flow_ID": "Flow_ID",  # No change needed
        "Flow_Data_Year": "Flow_Data_Year",  # No change needed
        "Flow_Material": "Flow_Material",  # No change needed
        "WC?": "WC?",  # No change needed
        "DM?": "DM?",  # No change needed
        "CC?": "CC?",  # No change needed
    },
    "2_1_Definition_Processes": {
        "ID": "ID",  # No change needed
        "Process_Name": "Process_Name",  # No change needed
        "Process_Logic": "Process_Logic",  # No change needed
        "TC_Configuration": "TC_Configuration",  # No change needed
        "Stock_Configuration": "Stock_Configuration",  # No change needed
        "Complete?": "Complete?",  # No change needed
        "Process_ID": "Process_ID",  # No change needed
    },
    "2_3_static_TCs": {
        "Flow_ID": "Flow_ID",  # No change needed
        "Process_ID": "Process_ID",  # No change needed
        "TC_material_ID": "TC_material_ID",  # No change needed
        "TC_Value_material": "TC_Value_material",  # No change needed
        "Complete?": "Complete?",  # No change needed
    },
    "2_4_dynamic_TCs": {
        "TC_material_ID": "TC_material_ID",  # No change needed
        "TC_Value_material": "TC_Value_material",  # No change needed
        "Year": "Year",  # No change needed
        "Process_ID": "Process_ID",  # No change needed
    },
    "2_5_Initial_Stock": {
        "Process_ID": "Process_ID",  # No change needed
        "Initial_Stock_material": "Initial_Stock_material",  # No change needed
        "Initial_Stock_WC[%]": "Initial_Stock_WC[%]",  # No change needed
        "Initial_Stock_DM[%]": "Initial_Stock_DM[%]",  # No change needed
        "Initial_Stock_CC[%]": "Initial_Stock_CC[%]",  # No change needed
        "Complete?": "Complete?",  # No change needed
    },
    "3_1_Definition_DSM": {
        "Process_ID": "Process_ID",  # No change needed
        "Process_Name": "Process_Name",  # No change needed
        "Process_Logic": "Process_Logic",  # No change needed
        "TC_Configuration": "TC_Configuration",  # No change needed
        "Stock_Configuration": "Stock_Configuration",  # No change needed
    },
    "3_2_Definition_FOMP": {
        "Process_ID": "Process_ID",  # No change needed
        "Process_Name": "Process_Name",  # No change needed
        "Process_Logic": "Process_Logic",  # No change needed
        "TC_Configuration": "TC_Configuration",  # No change needed
        "Stock_Configuration": "Stock_Configuration",  # No change needed
        "Complete?": "Complete?",  # No change needed
    },
    "4_1_Uncertainty_Parameters": {
        "Parameter_Name": "Parameter_Name",  # No change needed
        "Distribution": "Distribution",  # No change needed
        "Min": "Min",  # No change needed
        "Max": "Max",  # No change needed
        "Mean": "Mean",  # No change needed
        "StdDev": "StdDev",  # No change needed
        "Mode": "Mode",  # No change needed
    },
    "5_1_Scenario_Manager": {
        "Scenario_Name": "Scenario_Name",  # No change needed
        "Parameter_Name": "Parameter_Name",  # No change needed
        "start_year": "start_year",  # No change needed
        "end_year": "end_year",  # No change needed
    }
}

print("CRITICAL COLUMNS THAT NEED UPDATES:")
print("-" * 50)

# Check which columns actually need changes
changes_needed = []

for sheet, mappings in critical_mappings.items():
    if sheet in current_columns:
        for old_name, new_name in mappings.items():
            if old_name in current_columns[sheet] and old_name != new_name:
                changes_needed.append((sheet, old_name, new_name))
                print(f"  {sheet}: '{old_name}' → '{new_name}'")

if not changes_needed:
    print("  No critical column name changes needed!")
else:
    print(f"\nTotal critical changes needed: {len(changes_needed)}")

print("\n\nSHEET NAMING CHANGES NEEDED:")
print("-" * 50)

# Check sheet names
sheet_changes = []
current_sheets = list(wb.sheetnames)

if "PX - Template" in current_sheets:
    sheet_changes.append(("PX - Template", "PX_Template"))
    print(f"  'PX - Template' → 'PX_Template'")

if "CL_Visualisation" in current_sheets:
    sheet_changes.append(("CL_Visualisation", "CL_Visualization"))
    print(f"  'CL_Visualisation' → 'CL_Visualization'")

if not sheet_changes:
    print("  No sheet name changes needed!")

print(f"\nTotal sheet changes needed: {len(sheet_changes)}")

print("\n\nSUMMARY:")
print("="*80)
print(f"Critical column changes: {len(changes_needed)}")
print(f"Sheet name changes: {len(sheet_changes)}")
print(f"Total changes needed: {len(changes_needed) + len(sheet_changes)}")

if len(changes_needed) + len(sheet_changes) == 0:
    print("\n✅ NO CHANGES NEEDED - All naming is already correct!")
else:
    print(f"\n🔧 CHANGES REQUIRED - {len(changes_needed) + len(sheet_changes)} total changes needed")
